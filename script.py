# pip install beautifulsoup4 openpyxl fake_useragent selenium
# Race Meetings.xlsx

import re
import time
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
from fake_useragent import UserAgent
from openpyxl import load_workbook
from webdriver_manager.chrome import ChromeDriverManager

ua = UserAgent()
USER_AGENT = ua.random
# ChromeDriverPath = "C:/chromedriver/chromedriver.exe"
ChromeDriverPath = "/usr/local/bin/chromedriver"

BASE_URL = 'https://www.tab.com.au'
FILE_NAME = 'Race Meetings.xlsm'
target_column = 23
ALLOWED_MEETINGS = ['(VIC)', '(NSW)', '(QLD)', '(SA)', '(WA)', '(NT)', '(TAS)', '(ACT)', '(NZ)', '(NZL)']
FS = {}
SR = {}

# def setup_driver():
#     options = Options()
#     options.headless = True
#     options.add_argument("--disable-images")
#     options.add_argument("--start-maximized")
#     options.add_argument("--disable-popup-blocking")
#     options.add_argument(f"--user-agent={USER_AGENT}")
#     options.add_argument("--disable-blink-features=AutomationControlled")
#     options.add_argument("--disable-dev-shm-usage")
#     options.add_argument("--disable-gpu")
#     options.add_argument("--no-sandbox")
#     options.add_argument("--no-first-run")

#     service = Service(ChromeDriverPath)
#     driver = webdriver.Chrome(service=service, options=options)
#     driver.set_page_load_timeout(500)
#     driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    
#     return driver

def setup_driver():
    from webdriver_manager.chrome import ChromeDriverManager
    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("--disable-images")
    options.add_argument("--start-maximized")
    options.add_argument("--disable-popup-blocking")
    options.add_argument(f"--user-agent={USER_AGENT}")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--no-first-run")

    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)

    driver.set_page_load_timeout(500)
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    return driver



def find_all_races(html):
    soup = BeautifulSoup(html, 'html.parser')
    meetings = soup.find_all('div', {'data-testid': 'meeting', 'class': '_1e6ktkp'})
    race_links = soup.find_all('a', {'data-testid': 'race'})
    meetings_pre = [meeting.text for meeting in meetings]
    print(meetings_pre)
    meetings_names = []
    for meeting in meetings_pre:
        for allow in ALLOWED_MEETINGS:
            if allow.lower() in meeting.lower():
                meetings_names.append(meeting.split('(')[0].strip().lower())
    rounds_links = [link['href'] for link in race_links]

    print(meetings_names)

    return meetings_names, rounds_links

def extract_sky_rating(driver, url, meetings_names):
    global SR
    meeting_name = url.split('/')[3]

    if meeting_name.lower().replace('-', ' ') in meetings_names:
        SR.setdefault(meeting_name, {})

        try:
            driver.get(BASE_URL + url)
        except:
            driver.execute_script("window.stop()")

        time.sleep(2)
        html = driver.page_source
        soup = BeautifulSoup(html, "html.parser")

        # Each horse row
        rows = soup.select("div.row")

        for row in rows:
            try:
                horse_name = row.select_one("div.runner-name").get_text(strip=True)
                horse_name = horse_name.split("(")[0].strip()

                # Sky Rating is inside a <div> with a numeric value
                sr_column = row.find("div", string=re.compile(r"^\d+$"))
                if sr_column:
                    sky_rating = sr_column.get_text(strip=True)
                    SR[meeting_name][horse_name] = sky_rating
                    print("Sky:", meeting_name, horse_name, sky_rating)

            except Exception:
                continue

def extract_FS(driver, url, meetings_names):
    global FS
    meeting_name = url.split('/')[3]
    if meeting_name.lower().replace('-', ' ') in meetings_names:
        try:
            driver.get(BASE_URL + url)
        except:
            driver.execute_script("window.stop()")
        try:
            if FS[meeting_name]:
                pass
        except:
            FS[meeting_name] = {}
        try:
            button = WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, "//button[.//span[text()='Show All Form']]")))
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", button)
            button.click()
        except:
            pass
        while True:
            time.sleep(1)
            html = driver.page_source
            soup = BeautifulSoup(html, 'html.parser')
            FS_element = soup.find_all('p', {'class': 'comment-paragraph'})
            horse_name_divs = soup.find_all('div', {'class': 'row active'})
            if FS_element.__len__() > 0:
                print(FS_element.__len__(), horse_name_divs.__len__())
                for i in range(FS_element.__len__()):
                    horse_name = BeautifulSoup(str(horse_name_divs[i]), 'html.parser').find('div', {'class': 'runner-name'}).text.split('(')[0].strip()
                    FS[meeting_name][horse_name] = re.search(r"\(([-+]?\d*\.?\d+)\)", FS_element[i].text).group(1)
                print(FS)
                break


def get_meetings(driver, url):
    try:
        driver.get(url, )
    except:
        driver.execute_script("window.stop()")
    try:
        driver.get(url + 'R', )
    except:
        driver.execute_script("window.stop()")

    html = driver.page_source
    meetings_names, rounds_links = find_all_races(html=html)

    for i in range(rounds_links.__len__()):
        extract_FS(driver, rounds_links[i], meetings_names)
        extract_sky_rating(driver, rounds_links[i], meetings_names)
    

def merge_excel(excel_file, FS):
    workbook = load_workbook(filename=excel_file, keep_vba=True)

    def normalize(name):
        return name.strip().lower().replace("-", " ")
    normalized_sheet_map = {normalize(name): name for name in workbook.sheetnames}

    # Loop through sample_data and find best matching sheet
    # Insert Sky Rating into Column 24 (adjust as needed)
    sky_rating_column = 24  

    for raw_sheet_name, horses in SR.items():
        norm_name = normalize(raw_sheet_name)
        actual_sheet_name = normalized_sheet_map.get(norm_name)

        if actual_sheet_name:
            sheet = workbook[actual_sheet_name]
            print(f"Writing Sky Ratings into sheet: {actual_sheet_name}")

            for row in sheet.iter_rows(min_row=1):
                for cell in row:
                    horse_name = str(cell.value).strip() if cell.value else ""

                    if horse_name in horses:
                        sky = horses[horse_name]
                        sheet.cell(row=cell.row, column=sky_rating_column, value=sky)
                        print(f"Inserted Sky Rating {sky} for '{horse_name}' in row {cell.row}")
        else:
            print(f"No matching sheet found for '{raw_sheet_name}'")

    workbook.save(filename=excel_file)


def main():
    
    driver = setup_driver()
    get_meetings(driver=driver, url=BASE_URL + "/racing/meetings/today/")

    merge_excel(FILE_NAME, FS)

    driver.quit()


if __name__ == '__main__':
    main()
